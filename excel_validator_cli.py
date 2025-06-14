#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path
import re
from datetime import datetime
import json
from typing import Dict, List, Tuple, Optional
import sys
import argparse
import os


class StudentReportValidatorCLI:
    def __init__(self):
        self.validation_results = []
        
    def validate_file(self, file_path: str, check_scores=True, check_text_length=True, 
                     check_spelling=True, check_content=True):
        if not Path(file_path).exists():
            print(f"エラー: ファイルが見つかりません: {file_path}")
            return False
            
        try:
            self.validation_results = []
            
            # Load Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            print(f"\nファイルを検証中: {file_path}")
            print("=" * 80)
            
            # Perform validations
            if check_scores:
                print("\nテストスコアを検証中...")
                self.validate_test_scores(sheet)
                
            if check_text_length:
                print("文章の長さを検証中...")
                self.validate_text_sections(sheet)
                
            if check_spelling:
                print("誤字脱字をチェック中...")
                self.check_spelling_errors(sheet)
                
            if check_content:
                print("内容の適切性を検証中...")
                self.validate_content_appropriateness(sheet)
                
            wb.close()
            
            # Display results
            self.display_results()
            
            return True
            
        except Exception as e:
            print(f"エラー: ファイルの読み込み中にエラーが発生しました:\n{str(e)}")
            return False
            
    def validate_test_scores(self, sheet):
        score_cells = {
            '国語_目標': (10, 3),
            '社会_目標': (10, 4),
            '数学_目標': (10, 5),
            '理科_目標': (10, 6),
            '英語_目標': (10, 7),
            '合計_目標': (10, 8),
            '国語_結果': (12, 3),
            '社会_結果': (12, 4),
            '数学_結果': (12, 5),
            '理科_結果': (12, 6),
            '英語_結果': (12, 7),
            '合計_結果': (12, 8),
            '順位': (12, 9),
            '国語_平均': (14, 3),
            '社会_平均': (14, 4),
            '数学_平均': (14, 5),
            '理科_平均': (14, 6),
            '英語_平均': (14, 7),
            '合計_平均': (14, 8)
        }
        
        for label, (row, col) in score_cells.items():
            value = sheet.cell(row=row, column=col).value
            
            if value is None:
                self.add_validation_result(
                    f"テストスコア - {label}",
                    "未入力",
                    "警告",
                    f"{label}が入力されていません"
                )
            elif isinstance(value, (int, float)):
                if '合計' in label and '平均' not in label:
                    if value > 500:
                        self.add_validation_result(
                            f"テストスコア - {label}",
                            "範囲エラー",
                            "エラー",
                            f"{label}が異常に高い値です: {value}"
                        )
                elif '順位' in label:
                    if value < 1:
                        self.add_validation_result(
                            f"テストスコア - {label}",
                            "範囲エラー",
                            "エラー",
                            f"順位が無効な値です: {value}"
                        )
                else:
                    if value < 0 or value > 100:
                        self.add_validation_result(
                            f"テストスコア - {label}",
                            "範囲エラー",
                            "エラー",
                            f"{label}が0-100の範囲外です: {value}"
                        )
                        
        # Check if results match the sum
        subjects_result = []
        for col in range(3, 8):
            value = sheet.cell(row=12, column=col).value
            if value and isinstance(value, (int, float)):
                subjects_result.append(value)
                
        if len(subjects_result) == 5:
            calculated_sum = sum(subjects_result)
            total_result = sheet.cell(row=12, column=8).value
            if total_result and abs(calculated_sum - total_result) > 0.01:
                self.add_validation_result(
                    "テストスコア - 合計",
                    "計算エラー",
                    "エラー",
                    f"科目の合計({calculated_sum})と記載された合計({total_result})が一致しません"
                )
                
    def validate_text_sections(self, sheet):
        text_sections = {
            '現在の学習課題': (17, 2, 100, 500),
            '課題に対する進捗状況': (17, 10, 100, 500),
            '今後の目標': (27, 2, 50, 300),
            '目標に向けた指導計画': (27, 10, 100, 500),
            '授業態度・意欲・遅刻等': (37, 2, 50, 400),
            '宿題について': (37, 10, 50, 400),
            '家庭学習アドバイス': (47, 2, 100, 500),
            '夏期講習提案理由': (50, 2, 50, 300)
        }
        
        for section_name, (row, col, min_length, max_length) in text_sections.items():
            content = sheet.cell(row=row, column=col).value
            
            if not content:
                self.add_validation_result(
                    f"文章内容 - {section_name}",
                    "未入力",
                    "エラー",
                    f"{section_name}が入力されていません"
                )
            else:
                content_str = str(content).strip()
                content_length = len(content_str)
                
                if content_length < min_length:
                    self.add_validation_result(
                        f"文章内容 - {section_name}",
                        "文字数不足",
                        "警告",
                        f"{section_name}の文字数が少なすぎます（{content_length}文字、推奨: {min_length}文字以上）"
                    )
                elif content_length > max_length:
                    self.add_validation_result(
                        f"文章内容 - {section_name}",
                        "文字数超過",
                        "情報",
                        f"{section_name}の文字数が多すぎる可能性があります（{content_length}文字、推奨: {max_length}文字以下）"
                    )
                    
    def check_spelling_errors(self, sheet):
        common_typos = {
            'そして': ['そうして', 'そしして'],
            'ということ': ['とゆうこと', 'とゆう事'],
            '言う': ['ゆう'],
            'いう': ['ゆう'],
            'そういう': ['そうゆう'],
            'どういう': ['どうゆう'],
            '頑張': ['がんば'],
            '一生懸命': ['いっしょうけんめい', 'いっしょけんめい'],
            'できる': ['出来る'],
            'わかる': ['分かる', '判る'],
            'おこなう': ['行なう'],
            'あらわす': ['表わす', '現わす'],
        }
        
        text_cells = [(17, 2), (17, 10), (27, 2), (27, 10), (37, 2), (37, 10), (47, 2), (50, 2)]
        
        for row, col in text_cells:
            content = sheet.cell(row=row, column=col).value
            if content:
                content_str = str(content)
                
                # Check for common typos
                for correct, typos in common_typos.items():
                    for typo in typos:
                        if typo in content_str:
                            self.add_validation_result(
                                f"誤字脱字 - セル({row}, {col})",
                                "誤字の可能性",
                                "警告",
                                f"'{typo}' → '{correct}' の可能性があります"
                            )
                            
                # Check for repeated characters
                repeated_chars = re.findall(r'(.)\1{3,}', content_str)
                if repeated_chars:
                    self.add_validation_result(
                        f"誤字脱字 - セル({row}, {col})",
                        "文字の繰り返し",
                        "警告",
                        f"同じ文字の過度な繰り返しがあります: {''.join(repeated_chars)}"
                    )
                    
                # Check for missing punctuation
                if len(content_str) > 100 and content_str.count('。') < 2:
                    self.add_validation_result(
                        f"誤字脱字 - セル({row}, {col})",
                        "句読点不足",
                        "情報",
                        "長い文章に句読点が少ない可能性があります"
                    )
                    
    def validate_content_appropriateness(self, sheet):
        # 各セクションの詳細な検証ルール
        section_rules = {
            '現在の学習課題': {
                'keywords': ['課題', '問題', '苦手', '理解', '困難', '改善', '弱点', '不足'],
                'required_elements': ['具体的な科目や単元', '問題点の説明'],
                'negative_patterns': ['特になし', '問題なし', 'ありません'],
                'min_sentences': 2
            },
            '課題に対する進捗状況': {
                'keywords': ['進捗', '改善', '向上', '取り組み', '結果', '成果', '変化', '前回'],
                'required_elements': ['具体的な取り組み内容', '結果や変化'],
                'negative_patterns': ['変化なし', '進捗なし'],
                'min_sentences': 2
            },
            '今後の目標': {
                'keywords': ['目標', '点数', '成績', '内申', '向上', '達成', '点', '以上'],
                'required_elements': ['具体的な数値目標', '期限'],
                'negative_patterns': ['未定', '特になし'],
                'min_sentences': 1
            },
            '目標に向けた指導計画': {
                'keywords': ['指導', '計画', '方法', '実施', '授業', '学習', '対策', '強化'],
                'required_elements': ['具体的な指導方法', '実施頻度や期間'],
                'negative_patterns': ['継続', 'そのまま'],
                'min_sentences': 2
            },
            '授業態度・意欲・遅刻等': {
                'keywords': ['態度', '意欲', '集中', '参加', '遅刻', '欠席', '積極', '真面目'],
                'required_elements': ['態度の評価', '具体的な行動'],
                'negative_patterns': [],
                'min_sentences': 1
            },
            '宿題について': {
                'keywords': ['宿題', '提出', '取り組み', '正答', '理解度', '完成度', '期限', '質'],
                'required_elements': ['提出状況', '取り組みの質'],
                'negative_patterns': [],
                'min_sentences': 1
            },
            '家庭学習アドバイス': {
                'keywords': ['家庭', '学習', 'アドバイス', '方法', '時間', '習慣', '復習', '予習'],
                'required_elements': ['具体的な学習方法', '時間や頻度の提案'],
                'negative_patterns': ['特になし'],
                'min_sentences': 2
            }
        }
        
        row_col_map = {
            '現在の学習課題': (17, 2),
            '課題に対する進捗状況': (17, 10),
            '今後の目標': (27, 2),
            '目標に向けた指導計画': (27, 10),
            '授業態度・意欲・遅刻等': (37, 2),
            '宿題について': (37, 10),
            '家庭学習アドバイス': (47, 2)
        }
        
        for section_name, rules in section_rules.items():
            if section_name in row_col_map:
                row, col = row_col_map[section_name]
                content = sheet.cell(row=row, column=col).value
                
                if content:
                    content_str = str(content).strip()
                    content_lower = content_str.lower()
                    
                    # キーワードチェック
                    found_keywords = [kw for kw in rules['keywords'] if kw in content_lower]
                    keyword_ratio = len(found_keywords) / len(rules['keywords'])
                    
                    if keyword_ratio < 0.2:
                        self.add_validation_result(
                            f"内容確認 - {section_name}",
                            "キーワード不足",
                            "警告",
                            f"この項目に期待されるキーワードが不足しています。推奨キーワード: {', '.join(rules['keywords'][:4])}"
                        )
                    
                    # ネガティブパターンチェック
                    for pattern in rules.get('negative_patterns', []):
                        if pattern in content_lower:
                            self.add_validation_result(
                                f"内容確認 - {section_name}",
                                "不適切な表現",
                                "警告",
                                f"「{pattern}」という表現は避け、より具体的な内容を記載してください"
                            )
                    
                    # 文の数をチェック
                    sentences = re.split('[。！？]', content_str)
                    sentences = [s.strip() for s in sentences if s.strip()]
                    if len(sentences) < rules.get('min_sentences', 1):
                        self.add_validation_result(
                            f"内容確認 - {section_name}",
                            "文章構成",
                            "情報",
                            f"より詳細な説明が必要です（推奨: {rules.get('min_sentences', 1)}文以上）"
                        )
                    
                    # 具体性チェック
                    has_numbers = bool(re.search(r'\d+', content_str))
                    has_specific_terms = any(term in content_lower for term in ['具体的', '詳細', '例えば'])
                    
                    if section_name in ['今後の目標', '目標に向けた指導計画'] and not has_numbers:
                        self.add_validation_result(
                            f"内容確認 - {section_name}",
                            "具体性不足",
                            "情報",
                            "数値や具体的な期限を含めることを推奨します"
                        )
                        
    def add_validation_result(self, item: str, type: str, severity: str, detail: str):
        result = {
            'item': item,
            'type': type,
            'severity': severity,
            'detail': detail
        }
        self.validation_results.append(result)
        
    def display_results(self):
        if not self.validation_results:
            print("\n問題は見つかりませんでした。")
            return
            
        # Count by severity
        error_count = sum(1 for r in self.validation_results if r['severity'] == 'エラー')
        warning_count = sum(1 for r in self.validation_results if r['severity'] == '警告')
        info_count = sum(1 for r in self.validation_results if r['severity'] == '情報')
        
        print(f"\n\n検証結果サマリー:")
        print("=" * 80)
        print(f"エラー: {error_count}件")
        print(f"警告: {warning_count}件")
        print(f"情報: {info_count}件")
        print("=" * 80)
        
        # Display results by severity
        if error_count > 0:
            print("\n【エラー】")
            print("-" * 80)
            for result in self.validation_results:
                if result['severity'] == 'エラー':
                    print(f"項目: {result['item']}")
                    print(f"種類: {result['type']}")
                    print(f"詳細: {result['detail']}")
                    print("-" * 40)
                    
        if warning_count > 0:
            print("\n【警告】")
            print("-" * 80)
            for result in self.validation_results:
                if result['severity'] == '警告':
                    print(f"項目: {result['item']}")
                    print(f"種類: {result['type']}")
                    print(f"詳細: {result['detail']}")
                    print("-" * 40)
                    
        if info_count > 0:
            print("\n【情報】")
            print("-" * 80)
            for result in self.validation_results:
                if result['severity'] == '情報':
                    print(f"項目: {result['item']}")
                    print(f"種類: {result['type']}")
                    print(f"詳細: {result['detail']}")
                    print("-" * 40)
                    
    def save_report(self, output_file: str):
        if not self.validation_results:
            print("保存するチェック結果がありません")
            return
            
        try:
            if output_file.endswith('.json'):
                with open(output_file, 'w', encoding='utf-8') as f:
                    json.dump(self.validation_results, f, ensure_ascii=False, indent=2)
            else:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(f"生徒現状報告書チェック結果\n")
                    f.write(f"チェック日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}\n")
                    f.write("=" * 80 + "\n\n")
                    
                    for result in self.validation_results:
                        f.write(f"項目: {result['item']}\n")
                        f.write(f"種類: {result['type']}\n")
                        f.write(f"重要度: {result['severity']}\n")
                        f.write(f"詳細: {result['detail']}\n")
                        f.write("-" * 40 + "\n")
                        
            print(f"レポートを保存しました: {output_file}")
        except Exception as e:
            print(f"保存中にエラーが発生しました: {str(e)}")


def main():
    # 標準出力のエンコーディングを確実にUTF-8に設定
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    if sys.stderr.encoding != 'utf-8':
        sys.stderr.reconfigure(encoding='utf-8')
    
    # 環境変数でのエンコーディング設定
    os.environ.setdefault('PYTHONIOENCODING', 'utf-8')
    
    # ロケール設定の確認と警告
    import locale
    try:
        current_locale = locale.getlocale()
        preferred_encoding = locale.getpreferredencoding()
        if preferred_encoding.lower() not in ['utf-8', 'utf8']:
            print(f"警告: システムのエンコーディング設定が{preferred_encoding}です。")
            print("日本語が正しく表示されない場合は、以下のコマンドを実行してください:")
            print("export LANG=ja_JP.UTF-8")
            print("または")
            print("export LANG=C.UTF-8")
            print("=" * 60)
    except Exception as e:
        print(f"ロケール確認中にエラー: {e}")
    
    parser = argparse.ArgumentParser(description='生徒現状報告書チェッカー（コマンドライン版）')
    parser.add_argument('file', help='チェックするExcelファイルのパス')
    parser.add_argument('-o', '--output', help='結果を保存するファイル名（.txt or .json）')
    parser.add_argument('--no-scores', action='store_true', help='テストスコアのチェックをスキップ')
    parser.add_argument('--no-text', action='store_true', help='文章長のチェックをスキップ')
    parser.add_argument('--no-spelling', action='store_true', help='誤字脱字チェックをスキップ')
    parser.add_argument('--no-content', action='store_true', help='内容チェックをスキップ')
    
    args = parser.parse_args()
    
    validator = StudentReportValidatorCLI()
    
    success = validator.validate_file(
        args.file,
        check_scores=not args.no_scores,
        check_text_length=not args.no_text,
        check_spelling=not args.no_spelling,
        check_content=not args.no_content
    )
    
    if success and args.output:
        validator.save_report(args.output)
        
    return 0 if success else 1


if __name__ == "__main__":
    sys.exit(main())