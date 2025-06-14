# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox
from tkinterdnd2 import DND_FILES, TkinterDnD
import openpyxl
from pathlib import Path
import re
from datetime import datetime
import json
from typing import Dict, List, Tuple, Optional
import os
import sys


class StudentReportValidator:
    def __init__(self):
        # エンコーディング設定
        import locale
        import os
        
        # 環境変数でのエンコーディング設定
        os.environ.setdefault('PYTHONIOENCODING', 'utf-8')
        os.environ.setdefault('LANG', 'ja_JP.UTF-8')
        
        self.root = TkinterDnD.Tk()
        self.root.title("生徒現状報告書チェッカー")
        self.root.geometry("1200x800")
        
        # フォント設定（root window作成後）
        self.setup_fonts()
        
        # TkinterのUTF-8設定
        self.root.option_add('*Font', 'TkDefaultFont')
        
        self.current_file = None
        self.validation_results = []
        
        self.setup_ui()
        self.setup_drag_drop()
        
    def setup_fonts(self):
        """フォント設定（日本語対応）"""
        try:
            import tkinter.font as tkFont
            
            # 利用可能な日本語フォントを検索
            available_fonts = tkFont.families()
            japanese_fonts = [
                "Noto Sans CJK JP", "DejaVu Sans", "Liberation Sans", 
                "Takao Gothic", "VL Gothic", "IPAexGothic", "MS Gothic", 
                "Helvetica", "Arial", "TkDefaultFont"
            ]
            
            selected_font = "TkDefaultFont"
            for font in japanese_fonts:
                if font in available_fonts:
                    selected_font = font
                    break
            
            default_font = tkFont.nametofont("TkDefaultFont")
            default_font.configure(family=selected_font, size=10)
            
            text_font = tkFont.nametofont("TkTextFont")
            text_font.configure(family=selected_font, size=10)
            
            fixed_font = tkFont.nametofont("TkFixedFont")
            fixed_font.configure(family="Liberation Mono", size=10)
            
        except Exception as e:
            print(f"フォント設定中にエラー: {e}")
        
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # File selection frame with drag & drop area
        file_frame = ttk.LabelFrame(main_frame, text="ファイル選択", padding="10")
        file_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Create drag & drop area
        self.drop_frame = tk.Frame(file_frame, bg='lightgray', relief=tk.SUNKEN, bd=2)
        self.drop_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.drop_label = tk.Label(self.drop_frame, 
                                  text="ここにExcelファイルをドラッグ&ドロップ\nまたは下のボタンでファイルを選択",
                                  bg='lightgray', height=3)
        self.drop_label.pack(expand=True, fill='both', padx=20, pady=10)
        
        self.file_path_var = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.file_path_var, width=80).grid(row=1, column=0, padx=(0, 10))
        ttk.Button(file_frame, text="ファイルを選択", command=self.select_file).grid(row=1, column=1)
        ttk.Button(file_frame, text="チェック実行", command=self.validate_file).grid(row=1, column=2, padx=(10, 0))
        
        # Settings frame
        settings_frame = ttk.LabelFrame(main_frame, text="チェック設定", padding="10")
        settings_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.check_scores = tk.BooleanVar(value=True)
        self.check_text_length = tk.BooleanVar(value=True)
        self.check_spelling = tk.BooleanVar(value=True)
        self.check_content = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(settings_frame, text="テストスコアの妥当性", variable=self.check_scores).grid(row=0, column=0, sticky=tk.W)
        ttk.Checkbutton(settings_frame, text="文章量の適切性", variable=self.check_text_length).grid(row=0, column=1, sticky=tk.W)
        ttk.Checkbutton(settings_frame, text="誤字脱字チェック", variable=self.check_spelling).grid(row=1, column=0, sticky=tk.W)
        ttk.Checkbutton(settings_frame, text="内容の適切性", variable=self.check_content).grid(row=1, column=1, sticky=tk.W)
        
        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="チェック結果", padding="10")
        results_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        # Create Treeview for results
        columns = ('項目', '種類', '重要度', '詳細')
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show='tree headings', height=15)
        
        self.results_tree.heading('#0', text='')
        self.results_tree.heading('項目', text='項目')
        self.results_tree.heading('種類', text='種類')
        self.results_tree.heading('重要度', text='重要度')
        self.results_tree.heading('詳細', text='詳細')
        
        self.results_tree.column('#0', width=0, stretch=False)
        self.results_tree.column('項目', width=200)
        self.results_tree.column('種類', width=150)
        self.results_tree.column('重要度', width=100)
        self.results_tree.column('詳細', width=500)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=scrollbar.set)
        
        self.results_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # Summary frame
        summary_frame = ttk.Frame(main_frame)
        summary_frame.grid(row=3, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.summary_label = ttk.Label(summary_frame, text="チェック結果がここに表示されます")
        self.summary_label.grid(row=0, column=0, sticky=tk.W)
        
        ttk.Button(summary_frame, text="レポート保存", command=self.save_report).grid(row=0, column=1, sticky=tk.E, padx=(10, 0))
        
        # Configure tags for different severity levels
        self.results_tree.tag_configure('error', foreground='red')
        self.results_tree.tag_configure('warning', foreground='orange')
        self.results_tree.tag_configure('info', foreground='blue')
        self.results_tree.tag_configure('success', foreground='green')
        
    def setup_drag_drop(self):
        # Enable drag and drop on the drop frame
        self.drop_frame.drop_target_register(DND_FILES)
        self.drop_frame.dnd_bind('<<Drop>>', self.drop_file)
        
        # Visual feedback for drag over
        self.drop_frame.dnd_bind('<<DragEnter>>', self.drag_enter)
        self.drop_frame.dnd_bind('<<DragLeave>>', self.drag_leave)
        
    def drop_file(self, event):
        files = self.root.tk.splitlist(event.data)
        if files:
            file_path = files[0]
            if file_path.endswith(('.xlsx', '.xls')):
                self.file_path_var.set(file_path)
                self.drop_label.config(text=f"ファイルが選択されました:\n{Path(file_path).name}")
                # Auto-start validation
                self.root.after(100, self.validate_file)
            else:
                messagebox.showwarning("警告", "Excelファイル (.xlsx または .xls) を選択してください")
        self.drag_leave(event)
        
    def drag_enter(self, event):
        self.drop_frame.config(bg='lightblue')
        self.drop_label.config(bg='lightblue', text="ファイルをドロップしてください")
        
    def drag_leave(self, event):
        self.drop_frame.config(bg='lightgray')
        self.drop_label.config(bg='lightgray', text="ここにExcelファイルをドラッグ&ドロップ\nまたは下のボタンでファイルを選択")
        
    def select_file(self):
        filename = filedialog.askopenfilename(
            title="エクセルファイルを選択",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.file_path_var.set(filename)
            
    def validate_file(self):
        file_path = self.file_path_var.get()
        if not file_path:
            messagebox.showwarning("警告", "ファイルを選択してください")
            return
            
        if not Path(file_path).exists():
            messagebox.showerror("エラー", "ファイルが見つかりません")
            return
            
        try:
            self.validation_results = []
            self.results_tree.delete(*self.results_tree.get_children())
            
            # Load Excel file
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            
            # Perform validations
            if self.check_scores.get():
                self.validate_test_scores(sheet)
                
            if self.check_text_length.get():
                self.validate_text_sections(sheet)
                
            if self.check_spelling.get():
                self.check_spelling_errors(sheet)
                
            if self.check_content.get():
                self.validate_content_appropriateness(sheet)
                
            # Update summary
            self.update_summary()
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルの読み込み中にエラーが発生しました:\n{str(e)}")
            
    def validate_test_scores(self, sheet):
        score_cells = {
            '国語_目標': (10, 3),
            '社会_目標': (10, 4),
            '数学_目標': (10, 5),
            '理科_目標': (10, 6),
            '英語_目標': (10, 7),
            '合計_目標': (10, 8),
            '国語_結果': (12, 3),
            '社会_結果': (12, 4),
            '数学_結果': (12, 5),
            '理科_結果': (12, 6),
            '英語_結果': (12, 7),
            '合計_結果': (12, 8),
            '順位': (12, 9),
            '国語_平均': (14, 3),
            '社会_平均': (14, 4),
            '数学_平均': (14, 5),
            '理科_平均': (14, 6),
            '英語_平均': (14, 7),
            '合計_平均': (14, 8)
        }
        
        # Check if at least one subject score is entered for each category
        has_target_score = False
        has_result_score = False
        has_average_score = False
        
        for label, (row, col) in score_cells.items():
            value = sheet.cell(row=row, column=col).value
            
            # Track if any subject scores are entered
            if value is not None and isinstance(value, (int, float)):
                if '_目標' in label and '合計' not in label:
                    has_target_score = True
                elif '_結果' in label and '合計' not in label and '順位' not in label:
                    has_result_score = True
                elif '_平均' in label and '合計' not in label:
                    has_average_score = True
            
            # Only show warning for missing scores if NO scores are entered in that category
            if value is None:
                # Skip individual warnings - we'll check categories at the end
                continue
            elif isinstance(value, (int, float)):
                if '合計' in label and '平均' not in label:
                    if value > 500:
                        self.add_validation_result(
                            f"テストスコア - {label}",
                            "範囲エラー",
                            "エラー",
                            f"{label}が異常に高い値です: {value}"
                        )
                elif '順位' in label:
                    if value < 1:
                        self.add_validation_result(
                            f"テストスコア - {label}",
                            "範囲エラー",
                            "エラー",
                            f"順位が無効な値です: {value}"
                        )
                else:
                    if value < 0 or value > 100:
                        self.add_validation_result(
                            f"テストスコア - {label}",
                            "範囲エラー",
                            "エラー",
                            f"{label}が0-100の範囲外です: {value}"
                        )
        
        # Check if at least one score is entered in each category
        if not has_target_score:
            self.add_validation_result(
                "テストスコア - 目標点",
                "未入力",
                "警告",
                "目標点が1教科も入力されていません"
            )
        if not has_result_score:
            self.add_validation_result(
                "テストスコア - 結果",
                "未入力",
                "警告",
                "テスト結果が1教科も入力されていません"
            )
        if not has_average_score:
            self.add_validation_result(
                "テストスコア - 平均点",
                "未入力",
                "警告",
                "平均点が1教科も入力されていません"
            )
                        
        # Check if results match the sum
        subjects_result = []
        for col in range(3, 8):
            value = sheet.cell(row=12, column=col).value
            if value and isinstance(value, (int, float)):
                subjects_result.append(value)
                
        if len(subjects_result) >= 1:  # Changed from == 5 to >= 1
            calculated_sum = sum(subjects_result)
            total_result = sheet.cell(row=12, column=8).value
            if total_result and abs(calculated_sum - total_result) > 0.01:
                self.add_validation_result(
                    "テストスコア - 合計",
                    "計算エラー",
                    "エラー",
                    f"科目の合計({calculated_sum})と記載された合計({total_result})が一致しません"
                )
                
    def validate_text_sections(self, sheet):
        text_sections = {
            '現在の学習課題': (18, 2, 100, 500),
            '課題に対する進捗状況': (18, 10, 100, 500),
            '今後の目標': (28, 2, 50, 300),
            '目標に向けた指導計画': (28, 10, 100, 500),
            '授業態度・意欲・遅刻等': (38, 2, 50, 400),
            '宿題について': (38, 10, 50, 400),
            '家庭学習アドバイス': (48, 2, 100, 500),
            '夏期講習提案理由': (50, 2, 50, 300)
        }
        
        for section_name, (start_row, col, min_length, max_length) in text_sections.items():
            # Check all 4 rows in the content area
            all_content = []
            
            # Check 4 rows starting from start_row
            for row_offset in range(4):
                row = start_row + row_offset
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    all_content.append(str(cell_value).strip())
            
            # Combine all content found
            content = ' '.join(all_content) if all_content else None
            
            if not content:
                self.add_validation_result(
                    f"文章内容 - {section_name}",
                    "未入力",
                    "エラー",
                    f"{section_name}が入力されていません"
                )
            else:
                content_str = content.strip()
                content_length = len(content_str)
                
                if content_length < min_length:
                    self.add_validation_result(
                        f"文章内容 - {section_name}",
                        "文字数不足",
                        "警告",
                        f"{section_name}の文字数が少なすぎます（{content_length}文字、推奨: {min_length}文字以上）"
                    )
                elif content_length > max_length:
                    self.add_validation_result(
                        f"文章内容 - {section_name}",
                        "文字数超過",
                        "情報",
                        f"{section_name}の文字数が多すぎる可能性があります（{content_length}文字、推奨: {max_length}文字以下）"
                    )
                    
    def check_spelling_errors(self, sheet):
        common_typos = {
            'そして': ['そうして', 'そしして'],
            'ということ': ['とゆうこと', 'とゆう事'],
            '言う': ['ゆう'],
            'いう': ['ゆう'],
            'そういう': ['そうゆう'],
            'どういう': ['どうゆう'],
            '頑張': ['がんば'],
            '一生懸命': ['いっしょうけんめい', 'いっしょけんめい'],
            'できる': ['出来る'],
            'わかる': ['分かる', '判る'],
            'おこなう': ['行なう'],
            'あらわす': ['表わす', '現わす'],
        }
        
        text_cells = [(18, 2), (18, 10), (28, 2), (28, 10), (38, 2), (38, 10), (48, 2), (50, 2)]
        
        for start_row, col in text_cells:
            # Check all 4 rows in each content area
            all_content = []
            for row_offset in range(4):
                row = start_row + row_offset
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    all_content.append(str(cell_value).strip())
            
            # Combine all content found
            if all_content:
                content_str = ' '.join(all_content)
                
                # Check for common typos
                for correct, typos in common_typos.items():
                    for typo in typos:
                        if typo in content_str:
                            self.add_validation_result(
                                f"誤字脱字 - セル({start_row}, {col})",
                                "誤字の可能性",
                                "警告",
                                f"'{typo}' → '{correct}' の可能性があります"
                            )
                            
                # Check for repeated characters
                repeated_chars = re.findall(r'(.)\1{3,}', content_str)
                if repeated_chars:
                    self.add_validation_result(
                        f"誤字脱字 - セル({start_row}, {col})",
                        "文字の繰り返し",
                        "警告",
                        f"同じ文字の過度な繰り返しがあります: {''.join(repeated_chars)}"
                    )
                    
                # Check for missing punctuation
                if len(content_str) > 100 and content_str.count('。') < 2:
                    self.add_validation_result(
                        f"誤字脱字 - セル({start_row}, {col})",
                        "句読点不足",
                        "情報",
                        "長い文章に句読点が少ない可能性があります"
                    )
                    
    def validate_content_appropriateness(self, sheet):
        # 各セクションの詳細な検証ルール
        section_rules = {
            '現在の学習課題': {
                'keywords': ['課題', '問題', '苦手', '理解', '困難', '改善', '弱点', '不足'],
                'required_elements': ['具体的な科目や単元', '問題点の説明'],
                'negative_patterns': ['特になし', '問題なし', 'ありません'],
                'min_sentences': 2
            },
            '課題に対する進捗状況': {
                'keywords': ['進捗', '改善', '向上', '取り組み', '結果', '成果', '変化', '前回'],
                'required_elements': ['具体的な取り組み内容', '結果や変化'],
                'negative_patterns': ['変化なし', '進捗なし'],
                'min_sentences': 2
            },
            '今後の目標': {
                'keywords': ['目標', '点数', '成績', '内申', '向上', '達成', '点', '以上'],
                'required_elements': ['具体的な数値目標', '期限'],
                'negative_patterns': ['未定', '特になし'],
                'min_sentences': 1
            },
            '目標に向けた指導計画': {
                'keywords': ['指導', '計画', '方法', '実施', '授業', '学習', '対策', '強化'],
                'required_elements': ['具体的な指導方法', '実施頻度や期間'],
                'negative_patterns': ['継続', 'そのまま'],
                'min_sentences': 2
            },
            '授業態度・意欲・遅刻等': {
                'keywords': ['態度', '意欲', '集中', '参加', '遅刻', '欠席', '積極', '真面目'],
                'required_elements': ['態度の評価', '具体的な行動'],
                'negative_patterns': [],
                'min_sentences': 1
            },
            '宿題について': {
                'keywords': ['宿題', '提出', '取り組み', '正答', '理解度', '完成度', '期限', '質'],
                'required_elements': ['提出状況', '取り組みの質'],
                'negative_patterns': [],
                'min_sentences': 1
            },
            '家庭学習アドバイス': {
                'keywords': ['家庭', '学習', 'アドバイス', '方法', '時間', '習慣', '復習', '予習'],
                'required_elements': ['具体的な学習方法', '時間や頻度の提案'],
                'negative_patterns': ['特になし'],
                'min_sentences': 2
            }
        }
        
        row_col_map = {
            '現在の学習課題': (18, 2),
            '課題に対する進捗状況': (18, 10),
            '今後の目標': (28, 2),
            '目標に向けた指導計画': (28, 10),
            '授業態度・意欲・遅刻等': (38, 2),
            '宿題について': (38, 10),
            '家庭学習アドバイス': (48, 2)
        }
        
        for section_name, rules in section_rules.items():
            if section_name in row_col_map:
                start_row, col = row_col_map[section_name]
                
                # Check all 4 rows in the content area
                all_content = []
                for row_offset in range(4):
                    row = start_row + row_offset
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value:
                        all_content.append(str(cell_value).strip())
                
                # Combine all content found
                if all_content:
                    content_str = ' '.join(all_content)
                    content_lower = content_str.lower()
                    
                    # キーワードチェック
                    found_keywords = [kw for kw in rules['keywords'] if kw in content_lower]
                    keyword_ratio = len(found_keywords) / len(rules['keywords'])
                    
                    if keyword_ratio < 0.2:
                        self.add_validation_result(
                            f"内容確認 - {section_name}",
                            "キーワード不足",
                            "警告",
                            f"この項目に期待されるキーワードが不足しています。推奨キーワード: {', '.join(rules['keywords'][:4])}"
                        )
                    
                    # ネガティブパターンチェック
                    for pattern in rules.get('negative_patterns', []):
                        if pattern in content_lower:
                            self.add_validation_result(
                                f"内容確認 - {section_name}",
                                "不適切な表現",
                                "警告",
                                f"「{pattern}」という表現は避け、より具体的な内容を記載してください"
                            )
                    
                    # 文の数をチェック
                    sentences = re.split('[。！？]', content_str)
                    sentences = [s.strip() for s in sentences if s.strip()]
                    if len(sentences) < rules.get('min_sentences', 1):
                        self.add_validation_result(
                            f"内容確認 - {section_name}",
                            "文章構成",
                            "情報",
                            f"より詳細な説明が必要です（推奨: {rules.get('min_sentences', 1)}文以上）"
                        )
                    
                    # 具体性チェック
                    has_numbers = bool(re.search(r'\d+', content_str))
                    has_specific_terms = any(term in content_lower for term in ['具体的', '詳細', '例えば'])
                    
                    if section_name in ['今後の目標', '目標に向けた指導計画'] and not has_numbers:
                        self.add_validation_result(
                            f"内容確認 - {section_name}",
                            "具体性不足",
                            "情報",
                            "数値や具体的な期限を含めることを推奨します"
                        )
                        
    def add_validation_result(self, item: str, type: str, severity: str, detail: str):
        result = {
            'item': item,
            'type': type,
            'severity': severity,
            'detail': detail
        }
        self.validation_results.append(result)
        
        # Determine tag based on severity
        tag = 'info'
        if severity == 'エラー':
            tag = 'error'
        elif severity == '警告':
            tag = 'warning'
        elif severity == '成功':
            tag = 'success'
            
        # Add to tree
        self.results_tree.insert('', 'end', values=(item, type, severity, detail), tags=(tag,))
        
    def update_summary(self):
        error_count = sum(1 for r in self.validation_results if r['severity'] == 'エラー')
        warning_count = sum(1 for r in self.validation_results if r['severity'] == '警告')
        info_count = sum(1 for r in self.validation_results if r['severity'] == '情報')
        
        summary_text = f"チェック完了: エラー {error_count}件, 警告 {warning_count}件, 情報 {info_count}件"
        
        if error_count == 0 and warning_count == 0:
            summary_text += " - 重大な問題は見つかりませんでした。"
            self.add_validation_result(
                "総合評価",
                "チェック完了",
                "成功",
                "ファイルは概ね適切に作成されています"
            )
            
        self.summary_label.config(text=summary_text)
        
    def save_report(self):
        if not self.validation_results:
            messagebox.showinfo("情報", "保存するチェック結果がありません")
            return
            
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if filename:
            try:
                if filename.endswith('.json'):
                    with open(filename, 'w', encoding='utf-8') as f:
                        json.dump(self.validation_results, f, ensure_ascii=False, indent=2)
                else:
                    with open(filename, 'w', encoding='utf-8') as f:
                        f.write(f"生徒現状報告書チェック結果\n")
                        f.write(f"チェック日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}\n")
                        f.write(f"対象ファイル: {self.file_path_var.get()}\n")
                        f.write("=" * 80 + "\n\n")
                        
                        for result in self.validation_results:
                            f.write(f"項目: {result['item']}\n")
                            f.write(f"種類: {result['type']}\n")
                            f.write(f"重要度: {result['severity']}\n")
                            f.write(f"詳細: {result['detail']}\n")
                            f.write("-" * 40 + "\n")
                            
                messagebox.showinfo("成功", "レポートを保存しました")
            except Exception as e:
                messagebox.showerror("エラー", f"保存中にエラーが発生しました:\n{str(e)}")
                
    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    # エンコーディングの設定と確認
    import locale
    
    # 標準出力のエンコーディングを確実にUTF-8に設定
    if hasattr(sys.stdout, 'reconfigure'):
        if sys.stdout.encoding != 'utf-8':
            sys.stdout.reconfigure(encoding='utf-8')
        if sys.stderr.encoding != 'utf-8':
            sys.stderr.reconfigure(encoding='utf-8')
    
    # 環境変数でのエンコーディング設定
    os.environ.setdefault('PYTHONIOENCODING', 'utf-8')
    
    
    # ロケール設定の確認
    try:
        current_locale = locale.getlocale()
        preferred_encoding = locale.getpreferredencoding()
        if preferred_encoding.lower() not in ['utf-8', 'utf8']:
            print(f"警告: システムのエンコーディング設定が{preferred_encoding}です。")
            print("日本語が正しく表示されない場合は、以下のコマンドを実行してください:")
            print("export LANG=ja_JP.UTF-8 または export LANG=C.UTF-8")
    except Exception as e:
        print(f"ロケール確認中にエラー: {e}")
    
    app = StudentReportValidator()
    app.run()